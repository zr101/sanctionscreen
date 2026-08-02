"""Parser for the DFAT Consolidated Sanctions List (XLSX).

Format notes (new-format workbook published since Nov 2025):
- Single worksheet "Consolidated List", headers in row 1.
- The Reference column is like "2a", "2b", "3a": the numeric prefix groups
  all name rows belonging to one listed entity; each row is one name.
- Name Type is one of Primary Name / Alias / Original Script; Original
  Script rows hold Arabic/Cyrillic spellings and stay searchable.
- Date cells may come back from openpyxl as datetime objects or strings.
- "Control Date" is the closest structured field to a listing date
  (DECISIONS.md D8).
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sanctionscreen.ingestion.base import ParsedEntity, ParsedName

_REFERENCE_RE = re.compile(r"^(\d+)\s*([a-z]*)$", re.IGNORECASE)

_ENTITY_TYPES = {
    "individual": "individual",
    "entity": "entity",
    "vessel": "vessel",
    "aircraft": "aircraft",
}

_NAME_TYPES = {
    "primary name": "primary",
    "alias": "alias",
    "aka": "aka",
    "original script": "original_script",
}


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def parse_dfat(path: Path) -> list[ParsedEntity]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = (
        workbook["Consolidated List"]
        if "Consolidated List" in workbook.sheetnames
        else workbook.worksheets[0]
    )
    rows = sheet.iter_rows(values_only=True)
    header = [(_cell_str(h) or "").lower() for h in next(rows)]

    def col(row: tuple, *names: str) -> str | None:
        for name in names:
            if name in header:
                return _cell_str(row[header.index(name)])
        return None

    groups: dict[str, list[dict[str, str | None]]] = {}
    for raw in rows:
        reference = col(raw, "reference")
        name = col(raw, "name of individual or entity", "name")
        if not reference or not name:
            continue
        match = _REFERENCE_RE.match(reference)
        group_key = match.group(1) if match else reference
        groups.setdefault(group_key, []).append(
            {
                "reference": reference,
                "name": name,
                "type": col(raw, "type"),
                "name_type": col(raw, "name type"),
                "alias_strength": col(raw, "alias strength"),
                "date_of_birth": col(raw, "date of birth"),
                "place_of_birth": col(raw, "place of birth"),
                "citizenship": col(raw, "citizenship"),
                "address": col(raw, "address"),
                "additional_information": col(raw, "additional information"),
                "listing_information": col(raw, "listing information"),
                "committees": col(raw, "committees"),
                "control_date": col(raw, "control date"),
                "instrument": col(raw, "instrument of designation"),
            }
        )
    workbook.close()

    entities: list[ParsedEntity] = []
    for group_key, group_rows in groups.items():
        primary_row = next(
            (r for r in group_rows if (r["name_type"] or "").lower() == "primary name"),
            group_rows[0],
        )
        names = [
            ParsedName(
                name_type=_NAME_TYPES.get((r["name_type"] or "").lower(), "alias"),
                original=r["name"] or "",
                quality=r["alias_strength"],
            )
            for r in group_rows
        ]
        entities.append(
            ParsedEntity(
                source_list="DFAT",
                reference_number=group_key,
                primary_name=primary_row["name"] or "",
                entity_type=_ENTITY_TYPES.get((primary_row["type"] or "").lower(), "entity"),
                nationality=primary_row["citizenship"],
                date_of_birth=primary_row["date_of_birth"],
                listed_date=primary_row["control_date"],
                raw_record={"rows": group_rows},
                names=names,
            )
        )
    return entities
